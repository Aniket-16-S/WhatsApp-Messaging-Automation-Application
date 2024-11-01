import customtkinter as tk
from tkinter import messagebox
import openpyxl
import datetime
import time
import pyautogui as pgui
import sqlite3
import os
import sys
from tkinter import filedialog

file_path_global = ""
msg_global:str = None
open_delay:int = 0
msg_delay:float = 0






def validator(file_path):
    
    fp = str(file_path)
    new_p:str
    global file_path_global
    try :
        if fp[1] == '(' or fp[-1] == ')' or fp[-2] == ',' :
            new_p = fp[2:-3]
        else :
            new_p = fp
        file_path_global = new_p
       
    except IndexError as e :
        sys.exit()






def create_database(db_file):
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute('CREATE TABLE IF NOT EXISTS file_paths (file_path TEXT)')
    conn.commit()
    conn.close()

def get_file_path(db_file):
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute('SELECT file_path FROM file_paths')
    file_path = cursor.fetchone()
    conn.close()
    return file_path

def set_file_path(db_file, file_path):
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute('CREATE TABLE IF NOT EXISTS file_paths (file_path TEXT)')
    cursor.execute('INSERT INTO file_paths (file_path) VALUES (?)', (file_path,))
    conn.commit()
    conn.close()

def clear_everything():
    conn = sqlite3.connect("app_database.db")
    cursor = conn.cursor()
    cursor.execute("DELETE FROM sqlite_master WHERE type='table';")
    cursor.execute("DELETE FROM sqlite_master WHERE type='view';")
    cursor.execute("DELETE FROM sqlite_master WHERE type='index';")
    cursor.execute("DELETE FROM sqlite_master WHERE type='trigger';")
    cursor.execute("VACUUM;")
    conn.commit()
    conn.close()

def create_config_table(db_file):
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS config (
            message TEXT,
            msg_delay REAL,
            open_delay INTEGER
        )
    ''')
    conn.commit()
    conn.close()

def get_config(db_file):
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    cursor.execute('SELECT message, msg_delay, open_delay FROM config')
    config = cursor.fetchone()
    conn.close()
    return config

def set_config(db_file, message=None, msg_delay=None, open_delay=None):
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    create_config_table(db_file) 
    if message is None:
        message = None 
    if msg_delay is None:
        msg_delay = 1.8
    if open_delay is None:
        open_delay = 20
    cursor.execute('''
        INSERT INTO config (message, msg_delay, open_delay)
        VALUES (?, ?, ?)
        ON CONFLICT DO UPDATE SET
            message = excluded.message,
            msg_delay = excluded.msg_delay,
            open_delay = excluded.open_delay
    ''', (message, msg_delay, open_delay))
    conn.commit()
    conn.close()



def change_config(db_file, message=None, msg_delay=None, open_delay=None):
    conn = sqlite3.connect(db_file)
    cursor = conn.cursor()
    create_config_table(db_file)  
    update_values = []
    update_sql = []

    if message is not None:
        update_values.append(message)
        update_sql.append("message = ?")

    if msg_delay is not None:
        update_values.append(msg_delay)
        update_sql.append("msg_delay = ?")

    if open_delay is not None:
        update_values.append(open_delay)
        update_sql.append("open_delay = ?")

    if update_sql:
        update_sql_str = ", ".join(update_sql)
        cursor.execute(f'''
            UPDATE config
            SET {update_sql_str}
            WHERE 1  
        ''', update_values)
        conn.commit()

    conn.close()



def save_image_path(db_file) :
    conn = sqlite3.connect(db_file)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS image_file (
                image_path TEXT
                )
                ''')
    
    conn.commit()

    conn.close()


def get_image_path(db_file) :
    conn = sqlite3.connect(db_file)
    c = conn.cursor()
    save_image_path(db_file)
    c.execute("SELECT image_path FROM image_file")
    path = c.fetchone()
    c.close()
    if path :
        return path[0]
    else :
        return None

def set_image_path(db_file, path=None):
    conn = sqlite3.connect(db_file)
    c = conn.cursor()
    save_image_path(db_file)
    try :
        c.execute('INSERT INTO image_file (image_path) VALUES (?)',(path,))
    except Exception as e:
        print(e)
    conn.commit()
    conn.close()
    
def update_image_path(db_file, current_path, new_path):
    conn = sqlite3.connect(db_file)
    c = conn.cursor()
    
    try:
        c.execute('UPDATE image_file SET image_path = ? WHERE image_path = ?', (new_path, current_path))
    except Exception as e:
        print(e)
    
    conn.commit()
    conn.close()




def valid_img_path(db_file):
    way = get_image_path(db_file)
    way = str(way)
    if way[-1] == ')' :
        way = way[2:-3]
    return way







class Application(tk.CTk):

    
    def __init__(self, master = None):
        super().__init__(master)
        self.master = master
        self.title("RedVelvet")
        self.geometry("1020x650")
        self.create_widgets()
        self.minsize(width=350, height=400)

    
    def create_widgets(self):

        self.main_menu_label = tk.CTkLabel(self, text="Welcome to RedVelvet", font=("Roboto", 22))
        self.main_menu_label.place(relx='0.5', rely='0.08', anchor="center")

        self.add_data_button = tk.CTkButton(self, text="Add new Costumer Data", command=self.store_information, corner_radius=10, border_width=2)

        self.add_data_button.place(relx='0.27', rely='0.2', anchor="center")

        self.delete_data_button = tk.CTkButton(self, text="Delete Costumer Data", command=self.delete_data, corner_radius=10, border_width=2)

        self.delete_data_button.place(relx='0.27', rely='0.4', anchor="center")

        self.Edit_msg_button = tk.CTkButton(self, text=" View / Edit Message ", command=self.edit_msg, corner_radius=10, border_width=2 )

        self.Edit_msg_button.place(relx='0.27', rely='0.6', anchor="center")

        self.view_data_button = tk.CTkButton(self, text=" View Database ", command=self.print_information, corner_radius=10, border_width=2)

        self.view_data_button.place(relx='0.27', rely='0.8', anchor="center")       

        self.send_messages_button = tk.CTkButton(self, text="Check Birthdays & Send Messages", command=self.send_messages, corner_radius=10, border_width=2)

        self.send_messages_button.place(relx='0.73', rely='0.2', anchor="center")

        self.send_for_all_button = tk.CTkButton(self, text=" Send messages for all ", command=self.sendall, corner_radius=10, border_width=2)

        self.send_for_all_button.place(relx='0.73', rely='0.4', anchor="center")

        self.settings_button = tk.CTkButton(self, text="Settings", command=self.open_settings, corner_radius=7)
        self.settings_button.place(relx='0.5', rely='0.94', anchor='center')

        self.show_button = tk.CTkButton(self, text="Show current image", command=self.show_img, corner_radius=10, border_width=2)
        self.show_button.place(relx='0.73', rely='0.6', anchor="center")

        self.Img_button = tk.CTkButton(self, text="Select Image to share", command=self.img_window, border_width=2, corner_radius=10)
        self.Img_button.place(relx='0.73', rely='0.8', anchor="center")

        self.msg = msg_global
        self.delay_for_msg = msg_delay
        self.delay_to_open = open_delay
        self.db_file = 'app_database.db'

        x_px, y_px = pgui.size()
        self.x_t = 0.4568*x_px
        self.y_t = 0.2226*y_px

        self.x_cod = round(self.x_t)
        self.y_cod = round(self.y_t)

        
        self.img_file = valid_img_path(self.db_file)


    def img_window(self) :
        
        filepath = filedialog.askopenfilename(title="Select Image to copy", filetypes=[("Image files", ".jpg .jpeg .png .bmp")])

        if filepath :
            filepath = str(filepath)
            corrected_path:str = ""
            for ch in filepath:
                if ch == "/" :
                    for i in range(0,1) :
                        corrected_path += "\\"
                else :
                    corrected_path += ch

            if self.img_file == "None" :
                set_image_path(self.db_file, corrected_path)
            else :
                update_image_path(self.db_file, self.img_file, corrected_path)
            self.img_file = corrected_path

        return

    def show_img(self) :
        os.startfile(self.img_file)
        return
        


    def set_new_path(self):
        response = messagebox.askyesno("Factory Reset", "Are you sure you want to Reset to default ( Your data will NOT be deleted )?")
        if response:
            self.destroy()
            try:
                os.remove("app_database.db")
            except FileNotFoundError:
                clear_everything()
            app = App()
            app.mainloop()
        else :
            pass

    def contact(settings_window):
        contact_win = tk.CTkToplevel(settings_window.master)
        contact_win.title("Support")
        contact_win.geometry("300x200")
        contact_win.grab_set()

        lab = tk.CTkLabel(contact_win, text="Contact developer at \n\n  r230302@famt.ac.in \n \n to report any bugs or issues", font=("Roboto", 18))
        lab.pack(pady='20')


    def dest(self):
            self.destroy()

    def open_settings(self):
        settings_window = tk.CTkToplevel(self.master)
        settings_window.title("Settings")
        settings_window.geometry("320x300")
        settings_window.grab_set()

        button1 = tk.CTkButton(settings_window, text=" Factory Reset ", command=self.set_new_path) 
        button1.pack(pady=20)

        button_half = tk.CTkButton(settings_window, text="Contact Support", command=self.contact)
        button_half.pack(pady=20)

        button2 = tk.CTkButton(settings_window, text="Terms and Conditions", command=lambda sw=settings_window: self.show_terms_conditions(sw))
        button2.pack(pady=20)

        button3 = tk.CTkButton(settings_window, text="Change delays", command=lambda sw=settings_window: self.chnaging_delays(sw))
        button3.pack(pady=20)

        
    def chnaging_delays(self, settings_window) :
        Adv_SettingsWin = tk.CTkToplevel(settings_window.master)
        Adv_SettingsWin.title("Advance Settings")
        Adv_SettingsWin.geometry("350x300")
        Adv_SettingsWin.grab_set()

        general_delays_label = tk.CTkLabel(Adv_SettingsWin, text="Change general delays : ")
        general_delays_label.pack(pady=12)

        self.general_delays_var = tk.StringVar()
        self.general_delays_var.set(self.delay_for_msg)
        self.general_delays_var.trace("w", self.save_general_delays)
        general_delays_dropdown = tk.CTkOptionMenu(Adv_SettingsWin, variable=self.general_delays_var, values=[str(i) for i in [1.5, 1.6, 1.7, 1.8, 1.9, 2.0, 2.1, 2.2, 2.3, 2.4, 2.5, 2.6, 2.7, 2.8, 2.9, 3.0, 3.1, 3.2, 3.3, 3.4,  3.5, 3.6, 3.7, 3.8, 3.9, 4.0, 4.2, 4.3, 4.5]])

        general_delays_dropdown.pack(pady=12)


        open_delay_label = tk.CTkLabel(Adv_SettingsWin, text="Delay after whatsap Opens : ")
        open_delay_label.pack(pady=12)

        self.main_delay_var = tk.StringVar()
        self.main_delay_var.set(self.delay_to_open)
        self.main_delay_var.trace("w", self.save_main_delay)
        main_delay_dropdown = tk.CTkOptionMenu(Adv_SettingsWin, variable=self.main_delay_var, values=[str(i) for i in [15, 20, 25, 30, 35, 40, 45, 50, 60]])
        main_delay_dropdown.pack(pady=12)

        caution_label = tk.CTkLabel(Adv_SettingsWin, text="Change Above with caution! App's Performance may change.", text_color="red")
        caution_label.pack(pady=15)

        
    def save_main_delay(self, *args) :
        main_delay_value = self.main_delay_var.get()
        self.main_delay_var.set(main_delay_value)
        self.delay_to_open = main_delay_value
        change_config(self.db_file, open_delay=main_delay_value)
    
    def save_general_delays(self, *args):
        general_delays_value = self.general_delays_var.get()
        self.general_delays_var.set(general_delays_value)
        self.delay_for_msg = general_delays_value
        change_config(self.db_file, msg_delay=general_delays_value)
    
    def show_terms_conditions(self, settings_window):
        terms_window = tk.CTkToplevel(settings_window.master)
        terms_window.title("Terms and Conditions")
        terms_window.geometry("600x500")
        terms_window.grab_set()

        tcl = tk.CTkLabel(terms_window, text="Terms And Conditions", font=("Roboto", 20))
        tcl.pack(padx='5', pady='10')

        def ex():
            terms_window.destroy()
            settings_window.destroy()
            self.dest()
            

        def cl():
            terms_window.destroy()

        
        terms_text = tk.CTkTextbox(terms_window, wrap='word', width=500, height=350)
        terms_text.insert(tk.END, """ write your terms and conditions """)
        terms_text.pack(expand=True)

        acb = tk.CTkButton(terms_window, text="Accept", command=cl)
        rjb = tk.CTkButton(terms_window, text="Reject", command=ex)
        acb.place(relx='0.25', rely='0.95', anchor='center')
        rjb.place(relx='0.75', rely='0.95', anchor='center')


    def store_information(self):
        store_window = tk.CTkToplevel(self.master)
        store_window.title("Store Information")
        store_window.geometry("400x600")
        store_window.grab_set()

        name_label = tk.CTkLabel(store_window, text="Enter name:")
        name_label.pack(padx='0', pady='20')
        name_entry = tk.CTkEntry(store_window)
        name_entry.pack(padx='0', pady='20')

        phone_number_label = tk.CTkLabel(store_window, text="Enter phone number:")
        phone_number_label.pack(padx='0', pady='20')
        phone_number_entry = tk.CTkEntry(store_window)
        phone_number_entry.pack(padx='0', pady='20')

        birthdate_label = tk.CTkLabel(store_window, text="Enter birthdate (DD/MM):")
        birthdate_label.pack(padx='0', pady='20')
        birthdate_entry = tk.CTkEntry(store_window)
        birthdate_entry.pack(padx='0', pady='20')

        def save_data():
            name = name_entry.get()
            phone_number = phone_number_entry.get()
            birthdate = birthdate_entry.get()
            bd = str(birthdate)
            day = bd[0:2]
            if str(day)[-1] == "/" :
                day = day[:1]
                month = bd[2:4]
                if str(month)[-1] == "/" :

                    month = month[:1]

            else :
                month = bd[3:5]
                if str(month)[-1] == "/" :

                    month = month[:1]

            if int(day) > 31 or int(day) == 0  or int(month) > 12 or int(month) < 1 :
                    messagebox.showinfo("Value error", "You entered invalid birth-date. Please enter correct one.")
                    return
            if all(char.isdigit() for char in phone_number) and len(phone_number) == 10 : 
                workbook = openpyxl.load_workbook(file_path_global)
                sheet = workbook.active
                next_row = sheet.max_row + 1
                sheet.cell(row=next_row, column=1, value=name)
                sheet.cell(row=next_row, column=2, value=phone_number)
                sheet.cell(row=next_row, column=3, value=birthdate)
                workbook.save(file_path_global)
                messagebox.showinfo("Success", "Data saved successfully!")
                name_entry.delete(0, tk.END)
                phone_number_entry.delete(0, tk.END)
                birthdate_entry.delete(0, tk.END)
            else :
                messagebox.showinfo("Invalid Number Format", "An Invalid number was found, please check. Currently the app supports only Indian numbers. (regular 10 digits).")

        save_button = tk.CTkButton(store_window, text="Save", command=save_data)
        save_button.pack(padx='0', pady='20')



    def sendall(self):

        workbook = openpyxl.load_workbook(file_path_global)
        sheet = workbook.active

        phone_numbers = []
        names = []
        
        for row in sheet.iter_rows(min_row=2, values_only=True):
            name, phone_number, birthdate = row
            try :

                bd = str(birthdate)
                day = bd[0:2]
                if str(day)[-1] == "/" or str(day)[-1] == "-" :
                    day = day[:1]
                    month = bd[2:4]
                    if str(month)[-1] == "/" or str(month)[-1] == "-" :

                        month = month[:1]

                else :
                    month = bd[3:5]
                    if str(month)[-1] == "/" or str(month)[-1] == "-" :

                        month = month[:1]
                if int(day) > 31 or int(day) == 0  or int(month) > 12 or int(month) < 1 :
                    messagebox.showinfo("Value error", "Check sheet, You have added invalid date or month somewhere. Delete it.")
                    return
            except Exception as e:
                messagebox.showinfo("Value error", "Check sheet, You have added invalid date somewhere. Delete it.")
                return
                    
            phone_numbers.append(phone_number)
            names.append(name)

        if len(phone_numbers) >= 1 :
            pass
        else :
            messagebox.showinfo("Data not found ", "Data sheet is empty, Add few costumers")
            return

        if self.msg is None or self.msg == "" or self.msg == None:
            messagebox.showinfo("ValueError ", "Please edit todays message.")
            self.edit_msg()
            return
        response = messagebox.askyesno("Confirmation", "Messages will be sent for everyone. Sure?")
        if response:
            photo = messagebox.askyesno("Media file share", " Share the image ? ")
            try:
                if self.img_file and photo :
                    os.system(f"explorer /select,\"{self.img_file}\"")
                    time.sleep(1.5)
                    pgui.hotkey('ctrl', 'c')
                    time.sleep(0.8)
                pgui.press('win')
                time.sleep(self.delay_for_msg)
                pgui.typewrite('WhatsApp')
                time.sleep(self.delay_for_msg)
                pgui.press('enter')
                time.sleep(1.2)
                pgui.hotkey('win', 'up')
                time.sleep(3+self.delay_to_open)
            except pgui.FailSafeException:
                    messagebox.showinfo("Task Inturupted", "You Stopped the script ! (by moving mouse to corner of screen)")
                    return                
            except Exception as e:
                messagebox.showinfo("Unknown Inturrupt", "You Stopped the script ! (by moving mouse to corner of screen)")
                return

            for index, (name, phone_number) in enumerate(zip(names, phone_numbers)):
                try :
                    max_retries = 3
                    retries = 0                        
                    pgui.hotkey('ctrl', 'n')
                    time.sleep(1.2+self.delay_for_msg)
                    pgui.click(self.x_cod, self.y_cod, duration=0.3)
                    time.sleep(1.3+self.delay_for_msg)
                    time.sleep(1.3)
                    pgui.typewrite(phone_number)
                    time.sleep(1.4+self.delay_for_msg)
                    pgui.press('tab')
                    time.sleep(1.3)
                    pgui.press('enter')
                    time.sleep(1.4+self.delay_for_msg)
                    if self.img_file and photo :
                            pgui.hotkey('ctrl', 'v')
                            time.sleep(2)
                    for char in self.msg:
                        if char != '\n' :
                            pgui.typewrite(char, interval=0)
                        else :
                            pgui.hotkey('shift', 'enter')
                    time.sleep(self.delay_for_msg)
                    pgui.press('enter')
                    time.sleep(self.delay_for_msg)
                    pgui.hotkey('ctrl', 'w')
                    time.sleep(1)
                    pgui.press('esc')
                    time.sleep(1)
                    pgui.press('esc')
                    time.sleep(1.1)
                except pgui.FailSafeException:
                    messagebox.showinfo("Task Inturupted", "You Stopped the script ! (by moving mouse to corner of screen)")
                    return
                except Exception as e:
                    messagebox.showinfo("Unknown Inturupt", "Some isuues were detected (if this issue is occuring frequently, try factry reset from setting or contact developer)")
                    return
                        
                                           

            messagebox.showinfo("Task Complete", "Messages were sent to Everyone !")
                
        else:
            return



    def edit_msg(self):
        edit_window = tk.CTkToplevel(self.master)
        edit_window.title("Editing_Message")
        edit_window.geometry("450x600")
        edit_window.grab_set()

        msg_lable_c = tk.CTkLabel(edit_window, text="Last saved Message : ")
        msg_lable_c.pack(padx='0', pady='20')

        msg_c_text = tk.CTkTextbox(edit_window, wrap='word', height=130, width=400)
        msg_c_text.pack(padx='5', pady='20')
        
        msg_c_text.insert(tk.END, f"{self.msg}")
        
        msg_lable_new = tk.CTkLabel(edit_window, text="Enter new Message : ")
        msg_lable_new.pack(padx='0', pady='20')
        
        msg_entry = tk.CTkTextbox(edit_window, wrap='word', height=130, width=400)
        msg_entry.pack(padx='0', pady='20')

        msg_c_text.configure(state="disabled")

        def savedit():
            self.msg = msg_entry.get("1.0", "end-1c")
            change_config(self.db_file, message=self.msg)
            edit_window.grab_release()
            edit_window.destroy()
            messagebox.showinfo("Success", "Message saved successfully!")

        save_btn = tk.CTkButton(edit_window, text="Save", command=savedit)
        save_btn.pack(padx='0', pady='20')


    def delete_data(self):
        delete_window = tk.CTkToplevel(self.master)
        delete_window.title("Delete Data")
        delete_window.geometry("400x600")
        delete_window.grab_set()

        workbook = openpyxl.load_workbook(file_path_global)
        sheet = workbook.active
        data_label = tk.CTkLabel(delete_window, text="Select the row to delete:")
        data_label.pack(padx='0', pady='20')
        data_text = tk.CTkTextbox(delete_window, height=280, width=400, font=('Calibri', 17))
        data_text.pack(padx='0', pady='20')
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
            name, phone_number, birthdate = row
            if name is not None and phone_number is not None and birthdate is not None:
                data_text.insert(tk.END, f"{row_idx}. {name} \t  {phone_number} \t  {birthdate}\n")
        

        def delete_row():
            data_text.configure(state="normal")
            row_to_delete = int(row_entry.get())
            sheet.delete_rows(row_to_delete + 1)  # +1 because row indices start at 1 and not at 0
            workbook.save(file_path_global)
            messagebox.showinfo("Success", "Row deleted successfully!")
            data_text.delete(1.0, tk.END)
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=1):
                name, phone_number, birthdate = row
                if name is not None and phone_number is not None and birthdate is not None:
                    data_text.insert(tk.END, f"{row_idx}. {name}\t{phone_number}\t{birthdate}\n")
            row_entry.delete(0, tk.END)
            data_text.configure(state="disabled")

        row_label = tk.CTkLabel(delete_window, text="Enter the row number to delete:")
        row_label.pack(padx='0', pady='20')
        row_entry = tk.CTkEntry(delete_window)
        row_entry.pack(padx='0', pady='20')
        delete_button = tk.CTkButton(delete_window, text="Delete", command=delete_row)
        delete_button.pack(padx='0', pady='20')
        data_text.configure(state="disabled")

    def print_information(self):
        view_window = tk.CTkToplevel(master=self.master)
        view_window.title("View Data")
        view_window.geometry("400x450")
        view_window.grab_set()

        workbook = openpyxl.load_workbook(file_path_global)
        sheet = workbook.active
        data_label = tk.CTkLabel(view_window, text="Name \t        Phone Number \t      Birthdate")
        data_label.pack(padx='5', pady='20', anchor='w')
        data_text = tk.CTkTextbox(view_window, height=300, width=595, font=('Calibri', 18))
                
        data_text.pack(padx='5', pady='20', anchor='w')
        if sheet.max_row < 2:
            data_text.insert(tk.END, "The sheet is empty.")
        else:
            for row in sheet.iter_rows(min_row=2, values_only=True):
                name, phone_number, birthdate = row
                if name is not None and phone_number is not None and birthdate is not None:
                    data_text.insert(tk.END, f"{name} \t {phone_number} \t  {birthdate}\n" )
        data_text.configure(font=('Calibri', 18))

    def send_messages(self):
        
        workbook = openpyxl.load_workbook(file_path_global)
        sheet = workbook.active

        phone_numbers = []
        names = []
        today = datetime.date.today()
        for row in sheet.iter_rows(min_row=2, values_only=True):
            birth_date = None
            name, phone_number, birthdate = row
            try:
                birth_date = datetime.datetime.strptime(birthdate, "%d/%m/%y").date()
                
            except Exception as e:
                bd = str(birthdate)
                
                day = bd[0:2]
                if str(day)[-1] == "/" or str(day)[-1] == "-" :
                    day = day[:1]
                    month = bd[2:4]
                    if str(month)[-1] == "/" or str(month)[-1] == "-" :
                        
                        month = month[:1]
                        
                else :
                    month = bd[3:5]
                    if str(month)[-1] == "/" or str(month)[-1] == "-" :
                        
                        month = month[:1]
                       
                if int(day) > 31 or int(day) == 0  or int(month) > 12 or int(month) < 1 :
                    messagebox.showinfo("Value error", "Check sheet, You have added invalid date or month somewhere. Delete it.")
                    return
                
                if int(day) == int(today.day) and int(month) == int(today.month) :
                    phone_numbers.append(phone_number)
                    names.append(name)
                continue
                
            try:
                if (birth_date.month, birth_date.day) == (today.month, today.day):
                    phone_numbers.append(phone_number)
                    names.append(name)
            except Exception as e:
                pass

        if len(phone_numbers) >= 1 :
            pass
        else :
            messagebox.showinfo("No Birthdays ", "No Birthdays today !")
            return

        if self.msg is None or self.msg == "" or self.msg == None:
            messagebox.showinfo("ValueError ", "Please edit todays message.")
            self.edit_msg()
            return

        else:

            response = messagebox.askyesno("Assigning Task", "Please give confirmation to send messages. (by saying 'yes')")
            
            
            if response:
                wish_flag = messagebox.askyesno("Message Selection", "Add a Wish before message? ")
                photo = messagebox.askyesno("Media file share", " Share the image ? ")
                try:

                    if self.img_file and photo :
                        os.system(f"explorer /select,\"{self.img_file}\"")
                        time.sleep(1.5)
                        pgui.hotkey('ctrl', 'c')
                        time.sleep(0.8)                        
                    
                    pgui.press('win')
                    time.sleep(self.delay_for_msg)
                    pgui.typewrite('WhatsApp')
                    time.sleep(self.delay_for_msg)
                    pgui.press('enter')
                    time.sleep(1.2)
                    pgui.hotkey('win', 'up')
                    time.sleep(3+self.delay_to_open)
                    
                except pgui.FailSafeException:
                    messagebox.showinfo("Task Inturupted", "You Stopped the script ! (by moving mouse to corner of screen)")
                    return
                    

                except Exception as e:
                    messagebox.showinfo("Unknown error", "Some isuues were detected (if this issue is occuring frequently, try factry reset from setting or contact developer)")
                    #print("exp from send messages & says : ", e)
                    return

                for index, (name, phone_number) in enumerate(zip(names, phone_numbers)):
                    try :                        
                        pgui.hotkey('ctrl', 'n')
                        time.sleep(1.2+self.delay_for_msg)
                        #pgui.press('tab')                        
                        #time.sleep(1.3)
                        #pgui.press('tab')
                        pgui.click(self.x_cod, self.y_cod, duration=0.3)
                        time.sleep(1.3+self.delay_for_msg)
                        #pgui.press('enter')
                        time.sleep(1.3)
                        pgui.typewrite(phone_number)
                        time.sleep(1.4+self.delay_for_msg)
                        pgui.press('tab')
                        time.sleep(1.3)
                        pgui.press('enter')
                        time.sleep(1.4+self.delay_for_msg)
                        if self.img_file and photo :
                            pgui.hotkey('ctrl', 'v')
                            time.sleep(2)
                        
                        if len(name) >= 3 and wish_flag :
                            pgui.typewrite(f"Happy Birthday {name}! ")
                        for char in self.msg:
                            if char != '\n' :
                                pgui.typewrite(char, interval=0)
                            else :
                                pgui.hotkey('shift', 'enter')
                        
                        time.sleep(self.delay_for_msg)
                        pgui.press('enter')
                        time.sleep(self.delay_for_msg)
                        pgui.hotkey('ctrl', 'w')
                        time.sleep(1)
                        pgui.press('esc')
                        time.sleep(1)
                        pgui.press('esc')
                        time.sleep(1.1)
                    except pgui.FailSafeException:
                        messagebox.showinfo("Task Inturupted", "You Stopped the script ! (by moving mouse to corner of screen)")
                        return
                    except Exception as e:
                        messagebox.showinfo("Unknown error", "Some isuues were detected (if this issue is occuring frequently, try factry reset from setting or contact developer)")
                        return
                    
                                               

                messagebox.showinfo("Task Complete", "Messages were sent to Everyone !")
                
            else:
                return
                    

class App(tk.CTk):
    def __init__(self):
        super().__init__()
        self.geometry("500x300")
        self.title("Initial Setup")
        self.c_w()
        
    
    
    
    def c_w(self):
        self.grab_set()
        label = tk.CTkLabel(self, text="You will be prompted to select an Excel sheet that will be used to store information.", anchor="w")
        label_3 = tk.CTkLabel(self, text="If you don't have any,", anchor='w')
        label_4 = tk.CTkLabel(self, text="Create an empty Excel sheet ", anchor='w')
        label_2 = tk.CTkLabel(self, text="\nClick Select when ready")
        label.pack(padx='5', pady='10', anchor='w')
        label_3.pack(padx='5', pady='5', anchor='w')
        label_4.pack(padx='5', pady='5', anchor='w')
        label_2.pack(padx='5', pady='10', anchor='w')
        btn = tk.CTkButton(self, text="Select", command=self.close_win)
        btn.pack(padx='5', pady='10')
        
    def close_win(self):
        global file_path_global
        db_file = 'app_database.db'
        self.grab_release()
        self.destroy()
        file_path = tk.filedialog.askopenfilename(title='Select Excel File', filetypes=[('Excel Files', '*.xlsx')])
        validator(file_path)
        set_file_path(db_file, file_path_global)
        open_main_app()
            

def take_file_path():
    app = App()
    app.mainloop()
    

def open_main_app():
    app = Application()
    app.mainloop()


def main():
    db_file = 'app_database.db'
    create_database(db_file)
    create_config_table(db_file)
    save_image_path(db_file)
    file_path = get_file_path(db_file)
    my_config = get_config(db_file)
    file_path = str(file_path)
    check:str = 'None'
    check_2:str = '(None,)'
    c:str = "('',)"
    if my_config is None:
        set_config(db_file)
        my_config = get_config(db_file)  
    my_msg_delay = my_config[1]
    my_open_delay = my_config[2]
    my_message = my_config[0]
    global msg_global
    global open_delay
    global msg_delay
    try :
        msg_delay = my_msg_delay
        open_delay = my_open_delay
        msg_global = my_message
    except:
        msg_global = None
    import customtkinter as tk
    from tkinter import messagebox
    if file_path is None or file_path == check or file_path == check_2 or file_path == c :
        app = App()
        app.mainloop()
    else:
        validator(file_path)
        time.sleep(3)
        open_main_app()

    
if __name__ == '__main__':
    main()

